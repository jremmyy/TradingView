import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class StockSorter:
    def sort_stocks(self, stocks):
        try:
            # Sort by premarket volume (descending)
            sorted_stocks = sorted(stocks, key=lambda x: x.get('premarket_volume', 0), reverse=True)
            
            # Add rank for premarket volume
            for i, stock in enumerate(sorted_stocks, 1):
                stock['premarket_volume_rank'] = i
            
            # Sort by premarket change percentage (descending)
            sorted_stocks = sorted(sorted_stocks, key=lambda x: x.get('premarket_change_percent', 0), reverse=True)
            
            # Add rank for premarket change percentage
            for i, stock in enumerate(sorted_stocks, 1):
                stock['premarket_change_percent_rank'] = i
            
            # Sort by volume compared to average (descending)
            sorted_stocks = sorted(sorted_stocks, key=lambda x: x.get('volume', 0) / x.get('average_volume', 1) if x.get('average_volume', 0) != 0 else 0, reverse=True)
            
            # Add rank for volume compared to average
            for i, stock in enumerate(sorted_stocks, 1):
                stock['volume_vs_avg_rank'] = i
            
            # Calculate overall rank (lower is better)
            for stock in sorted_stocks:
                stock['overall_rank'] = (
                    stock.get('premarket_volume_rank', len(sorted_stocks)) +
                    stock.get('premarket_change_percent_rank', len(sorted_stocks)) +
                    stock.get('volume_vs_avg_rank', len(sorted_stocks))
                )
            
            # Sort by overall rank (ascending)
            return sorted(sorted_stocks, key=lambda x: x['overall_rank'])
        except Exception as e:
            logging.error(f"An error occurred while sorting stocks: {str(e)}")
            return stocks

    def export_to_excel(self, watchlists, filename='stock_rankings.xlsx'):
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove the default sheet
            
            # Create a sheet for overall rankings
            all_stocks = []
            for watchlist_name, stocks in watchlists.items():
                for stock in stocks:
                    stock['watchlist'] = watchlist_name
                all_stocks.extend(stocks)
            
            all_stocks_sorted = self.sort_stocks(all_stocks)
            df_all = pd.DataFrame(all_stocks_sorted)
            ws_all = wb.create_sheet("Overall Rankings")
            for r in dataframe_to_rows(df_all, index=False, header=True):
                ws_all.append(r)
            
            # Create a sheet for each watchlist
            for watchlist_name, stocks in watchlists.items():
                sorted_stocks = self.sort_stocks(stocks)
                df = pd.DataFrame(sorted_stocks)
                ws = wb.create_sheet(watchlist_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
            
            wb.save(filename)
            logging.info(f"Excel file '{filename}' has been created successfully.")
        except Exception as e:
            logging.error(f"An error occurred while exporting to Excel: {str(e)}")
